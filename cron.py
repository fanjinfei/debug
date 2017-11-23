# encoding: utf-8
import sys
import csv
import unicodecsv
import codecs
import traceback

import openpyxl
from crontab import CronTab

#format: [minute, hour, cmdline, comments]
weekday_jobs = [ 
]
daily_jobs = [
]

def write_xls(filename, sheets):
    #book = openpyxl.load_workbook('sheets.xlsx')
    book = openpyxl.Workbook()

    for sheet in sheets:
        ws = book.create_sheet(title=sheet.get('name', 'sheet 1'))
        for row in sheet.get('data',[]):
            ws.append(row)
        cols =  [col for col in ws.columns]
        widths = sheet.get('col_width', {})
        for k,v in widths.iteritems():
            ws.column_dimensions[cols[k][0].column].width = v
    try:
        sheet1 = book.get_sheet_by_name("Sheet")
        book.remove_sheet(sheet1)
    except:
        pass
    book.save(filename)


class CronData():
	def __init__(self, cron):
		self.cron = cron
	def help(self):
		print 'todo'
	def list(self):
		pass
	def diff(self):
		iter = self.cron.find_comment('test')
		jobs = [job for job in iter]
		if jobs:
			job = jobs[0]
			print 'found', job
			if True: #remove
				self.cron.remove(job)
				self.cron.write_to_user(user=True)
				print 'removed'
	def _data(self): #verify jobs
		pass
	def install(self):
		job = self.cron.new(command='touch /tmp/a.txt', comment='test')
		job.minute.every(1)
		job.enable()
		#job.every_reboot() run on reboot only
		#job.minute.during(5,50).every(5)
		#job.hour.every(4)
		#job.day.on(4, 5, 6)
		self.cron.write_to_user(user=True)
		print 'installed'
	def export(self):
            ''' rows.sort(key=lambda x:-x[3])
            rows.insert(0, ['ID / Identificateur',
                            'Title English / Titre en anglais',
                            'Title French / Titre en français',
                            'Number of downloads / Nombre de téléchargements'])
            rows.append(['total','','', org_stats.get(org_id)])
            sheets.append({'name':title,
                           'data': rows,
                           'col_width':{0:40, 1:50, 2:50, 3:40}
                           }
                          )
        sheets.sort(key=lambda x: x['name'])
        sheets.insert(0, sheet2)
        sheets.insert(0, sheet1)
        write_xls('/tmp/downloads.xls', sheets)'''

def main():
	system_cron   = CronTab()
	my_user_cron  = CronTab(user=True)
	#user_cron    = CronTab(user='jffan')

	for job in my_user_cron:
	    print job
	cront_data = CronData(my_user_cron)
	if len(sys.argv)==1:
		cront_data.help()
	elif sys.argv[1] == 'list':
		cront_data.list()
	elif sys.argv[1] == 'diff':
		cront_data.diff()
	elif sys.argv[1] == 'install':
		cront_data.install()
	elif sys.argv[1] == 'export':
		cront_data.export(*sys.argv[2:])



if __name__ =='__main__':
    main()
