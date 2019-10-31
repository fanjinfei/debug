#!/usr/bin/env python

# Copyright 2017 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.


#export GOOGLE_APPLICATION_CREDENTIALS=~/src/go/inbound-augury-252116-6e137952ff3b.json

"""DialogFlow API Detect Intent Python sample with text inputs.

Examples:
  python detect_intent_texts.py -h
  python detect_intent_texts.py --project-id PROJECT_ID \
  --session-id SESSION_ID \
  "hello" "book a meeting room" "Mountain View"
  python detect_intent_texts.py --project-id PROJECT_ID \
  --session-id SESSION_ID \
  "tomorrow" "10 AM" "2 hours" "10 people" "A" "yes"
  pgoogle_dialogflow_admin.py  --project-id PID  create test_1 \
    --training-phrases-parts "a test question" --message-texts "got you" 
"""
import openpyxl
import argparse
import uuid
import sys

def parse_args():
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(
        '--project-id',
        help='Project/agent id.  Required.',
        required=True)

    subparsers = parser.add_subparsers(dest='command')

    list_parser = subparsers.add_parser(
        'list', help=list_intents.__doc__)

    create_parser = subparsers.add_parser(
        'create', help=create_intent.__doc__)
    create_parser.add_argument(
        'display_name')
    create_parser.add_argument(
        '--training-phrases-parts',
        nargs='*',
        type=str,
        help='Training phrases.',
        default=[])
    create_parser.add_argument(
        '--message-texts',
        nargs='*',
        type=str,
        help='Message texts for the agent\'s response when the intent '
        'is detected.',
        default=[])

    update_parser = subparsers.add_parser(
        'update', help=update_intent.__doc__)
    update_parser.add_argument(
        'display_name')
    update_parser.add_argument(
        '--training-phrases-parts',
        nargs='*',
        type=str,
        help='Training phrases.',
        default=[])
    update_parser.add_argument(
        '--message-texts',
        nargs='*',
        type=str,
        help='Message texts for the agent\'s response when the intent '
        'is detected.',
        default=[])

    batch_update_parser = subparsers.add_parser(
        'batch-update', help=batch_update_intent.__doc__)
    batch_update_parser.add_argument(
        '--filename',
        nargs='*',
        type=str,
        help='intent xlsx file.',
        default=[])

    delete_parser = subparsers.add_parser(
        'delete', help=delete_intent.__doc__)
    delete_parser.add_argument(
        'intent_id',
        help='The id of the intent.')
        
    detect_parser = subparsers.add_parser(
        'detect', help=detect_intent_texts.__doc__)
    detect_parser.add_argument(
        '--session-id',
        help='Identifier of the DetectIntent session. '
        'Defaults to a random UUID.',
        default=str(uuid.uuid4()))
    detect_parser.add_argument(
        '--language-code',
        help='Language code of the query. Defaults to "en-US".',
        default='en-US')
    detect_parser.add_argument(
        'texts',
        nargs='+',
        type=str,
        help='Text inputs.')

    args = parser.parse_args()
    return args
        
    
# [START dialogflow_detect_intent_text]
def detect_intent_texts(project_id, session_id, texts, language_code):
    """Returns the result of detect intent with texts as inputs.

    Using the same `session_id` between requests allows continuation
    of the conversation."""
    import dialogflow_v2 as dialogflow
    session_client = dialogflow.SessionsClient()

    session = session_client.session_path(project_id, session_id)
    print('Session path: {}\n'.format(session))

    for text in texts:
        text_input = dialogflow.types.TextInput(
            text=text, language_code=language_code)

        query_input = dialogflow.types.QueryInput(text=text_input)

        response = session_client.detect_intent(
            session=session, query_input=query_input)

        print('=' * 20)
        print('Query text: {}'.format(response.query_result.query_text))
        print('Detected intent: {} (confidence: {})\n'.format(
            response.query_result.intent.display_name,
            response.query_result.intent_detection_confidence))
        print('Fulfillment text: {}\n'.format(
            response.query_result.fulfillment_text))
# [END dialogflow_detect_intent_text]

# [START dialogflow_list_intents]
def list_intents(project_id):
    import dialogflow_v2 as dialogflow
    
    intents_client = dialogflow.IntentsClient()

    parent = intents_client.project_agent_path(project_id)

    intents = intents_client.list_intents(parent, intent_view=1)

    for intent in intents:
        print('=' * 20)
        print('Intent name: {}'.format(intent.name))
        print('Intent display_name: {}'.format(intent.display_name))
        print('Action: {}\n'.format(intent.action))
        print('Root followup intent: {}'.format(
            intent.root_followup_intent_name))
        print('Parent followup intent: {}\n'.format(
            intent.parent_followup_intent_name))

        print('Input contexts:')
        for input_context_name in intent.input_context_names:
            print('\tName: {}'.format(input_context_name))

        print('Output contexts:')
        for output_context in intent.output_contexts:
            print('\tName: {}'.format(output_context.name))

        print('parameters:')
        for pa in intent.parameters:
            print('\tParameter name:{}, entity:{}'.format(pa.name, pa.value))
        print('Training phrases:', intent.training_phrases)
        for tr in intent.training_phrases:
            print('\t {}'.format(tr))
            for p in tr.parts:
                print('\t\t {}'.format(p))
# [END dialogflow_list_intents]

# [START dialogflow_create_intent]
def create_intent(project_id, display_name, training_phrases_parts,
                  message_texts, language_code="en", intents_client=None):
    """Create an intent of the given intent type."""
    import dialogflow_v2 as dialogflow
    intents_client = intents_client or dialogflow.IntentsClient()

    parent = intents_client.project_agent_path(project_id)
    training_phrases = []
    for training_phrases_part in training_phrases_parts:
        part = dialogflow.types.Intent.TrainingPhrase.Part(
            text=training_phrases_part)
        # Here we create a new training phrase for each provided part.
        training_phrase = dialogflow.types.Intent.TrainingPhrase(parts=[part])
        training_phrases.append(training_phrase)

    text = dialogflow.types.Intent.Message.Text(text=message_texts)
    message = dialogflow.types.Intent.Message(text=text)

    intent = dialogflow.types.Intent(
        display_name=display_name,
        training_phrases=training_phrases,
        messages=[message])

    response = intents_client.create_intent(parent, intent, language_code)

    print('Intent created: {}'.format(response))
# [END dialogflow_create_intent]

# [START dialogflow_update_intent]
def update_intent(project_id, display_name, training_phrases_parts,
                  message_texts,language_code='en'):
    """Update an intent of the given intent display_name."""
    import dialogflow_v2 as dialogflow
    intents_client = dialogflow.IntentsClient()

    parent = intents_client.project_agent_path(project_id)
    intents = intents_client.list_intents(parent)
    intent = None
    for intent in intents:
        if intent.display_name == display_name:
            break
    assert(intent)
    intent =  intents_client.get_intent(intent.name, intent_view=dialogflow.enums.IntentView.INTENT_VIEW_FULL)
    
    training_phrases = []
    for training_phrases_part in training_phrases_parts:
        part = dialogflow.types.Intent.TrainingPhrase.Part(
            text=training_phrases_part)
        # Here we create a new training phrase for each provided part.
        training_phrase = dialogflow.types.Intent.TrainingPhrase(parts=[part])
        training_phrases.append(training_phrase)

    text = dialogflow.types.Intent.Message.Text(text=message_texts)
    message = dialogflow.types.Intent.Message(text=text)

    # extend existing ones:
    #intent.training_phrases.extend(training_phrases)
    #intent.messages=[message] :: not working
    
    #create a new one
    intent = dialogflow.types.Intent(
        name=intent.name,
        display_name=display_name,
        training_phrases=training_phrases,
        messages=[message])

    response = intents_client.update_intent(intent, language_code)

    print('Intent updated: {}'.format(response))
# [END dialogflow_update_intent]

# [START dialogflow_batch_update_intent]
def batch_update_intent(project_id, filename, language_code='en'):
    """Batch update an intent of the given xls file."""
    import dialogflow_v2 as dialogflow
    res = read_xlsx(filename[0], "serviceDesk_faq")
    new_intents = []
    intent = None
    for row in res:
        if row[0] :
            if intent: 
                new_intents.append(intent)
            intent = {'name':row[0], 'training_phrases': [row[1]], 'message':row[2],'training_phrases_fr': [] }
        else:
            if row[1]:
                intent['training_phrases'].append(row[1])
        if row[3]: #french
            intent['training_phrases_fr'].append(row[3])
        if row[4]:
            intent['message_fr'] = row[4]
        

    if intent: new_intents.append(intent) #last one
    #print(intents)
    for intent in new_intents:
        #assume ALL has at least English version
        if intent.get('message_fr', None):
            assert( len(intent['training_phrases_fr']) >0 and intent.get('message', None) )

    intents_client = dialogflow.IntentsClient()

    parent = intents_client.project_agent_path(project_id)
    intents = intents_client.list_intents(parent)
    for intent in intents:
        for n_int in new_intents:
            if intent.display_name == n_int['name']:
                n_int['dialogflow_name'] = intent.name
    
    #update/create English version
    for n_int in new_intents:
        intent_name = n_int.get('dialogflow_name',None)
        if intent_name: #update
            intent =  intents_client.get_intent(intent_name, intent_view=dialogflow.enums.IntentView.INTENT_VIEW_FULL)
            training_phrases = []
            for training_phrases_part in n_int['training_phrases']:
                part = dialogflow.types.Intent.TrainingPhrase.Part(
                    text=training_phrases_part)
                # Here we create a new training phrase for each provided part.
                training_phrase = dialogflow.types.Intent.TrainingPhrase(parts=[part])
                training_phrases.append(training_phrase)
            text = dialogflow.types.Intent.Message.Text(text=[n_int['message']])
            message = dialogflow.types.Intent.Message(text=text)
            intent = dialogflow.types.Intent(
                name=intent_name,
                display_name=n_int['name'],
                training_phrases=training_phrases,
                messages=[message])

            response = intents_client.update_intent(intent, language_code)
            print('Intent updated: {}'.format(response))
        else: #create
            create_intent(project_id, n_int['name'], n_int['training_phrases'],
                  [n_int['message']], intents_client=intents_client)
    
    intents = intents_client.list_intents(parent)
    for intent in intents:
        for n_int in new_intents:
            if intent.display_name == n_int['name']:
                n_int['dialogflow_name'] = intent.name
    #Only update French version; they share same internal ID
    language_code = 'fr'
    for n_int in new_intents:
        if not n_int.get('message_fr', None): continue
        intent_name = n_int.get('dialogflow_name',None)
        assert(intent_name)

        intent =  intents_client.get_intent(intent_name, intent_view=dialogflow.enums.IntentView.INTENT_VIEW_FULL)
        training_phrases = []
        for training_phrases_part in n_int['training_phrases_fr']:
            part = dialogflow.types.Intent.TrainingPhrase.Part(
                text=training_phrases_part)
            # Here we create a new training phrase for each provided part.
            training_phrase = dialogflow.types.Intent.TrainingPhrase(parts=[part])
            training_phrases.append(training_phrase)
        text = dialogflow.types.Intent.Message.Text(text=[n_int['message_fr']])
        message = dialogflow.types.Intent.Message(text=text)
        intent = dialogflow.types.Intent(
            name=intent_name,
            display_name=n_int['name'],
            training_phrases=training_phrases,
            messages=[message])

        response = intents_client.update_intent(intent, language_code)
        print('Intent updated (fr): {}'.format(response))

# [END dialogflow_batch_update_intent]

# [START dialogflow_delete_intent]
def delete_intent(project_id, intent_id):
    """Delete intent with the given intent type and intent value."""
    import dialogflow_v2 as dialogflow
    intents_client = dialogflow.IntentsClient()

    intent_path = intents_client.intent_path(project_id, intent_id)

    intents_client.delete_intent(intent_path)
# [END dialogflow_delete_intent]


# Helper to get intent from display name.
def _get_intent_ids(project_id, display_name):
    import dialogflow_v2 as dialogflow
    intents_client = dialogflow.IntentsClient()

    parent = intents_client.project_agent_path(project_id)
    intents = intents_client.list_intents(parent)
    intent_names = [
        intent.name for intent in intents
        if intent.display_name == display_name]

    intent_ids = [
        intent_name.split('/')[-1] for intent_name
        in intent_names]

    return intent_ids
     
def main():
    args = parse_args()
    if args.command == 'list':
        list_intents(args.project_id)
    elif args.command == 'create':
        #print (args.message_texts)
        create_intent(
            args.project_id, args.display_name, args.training_phrases_parts,
            args.message_texts, )
    elif args.command == 'update':
        update_intent(
            args.project_id, args.display_name, args.training_phrases_parts,
            args.message_texts, )
    elif args.command == 'batch-update':
        batch_update_intent(
            args.project_id, args.filename)
    elif args.command == 'delete':
        delete_intent(args.project_id, args.intent_id)    
    elif args.command == 'detect':
        detect_intent_texts( args.project_id, args.session_id, args.texts, args.language_code)


def read_xlsx(filename, sheetname):
    print(filename, sheetname)
    book = openpyxl.load_workbook(filename)
    #book = openpyxl.Workbook()

    #print(book.sheetnames)
    sheet1 = book.get_sheet_by_name(sheetname)

    max_col = sheet1.max_column
    m_row = sheet1.max_row
    res = []
    for i in range(1, m_row + 1):
        row = []
        for j in range(1, max_col + 1):
            cell_obj = sheet1.cell(row = i, column = j) 
            print(cell_obj.value) 
            row.append(cell_obj.value)
        res.append(row)
    return res
        
if __name__ == '__main__':
    main()

